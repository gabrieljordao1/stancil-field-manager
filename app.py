import streamlit as st
import datetime
import json
import os
import copy
import re
import smtplib
import imaplib
import email as email_lib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import decode_header

# âââ PAGE CONFIG ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
st.set_page_config(page_title="Stancil Field Manager", page_icon="ð ", layout="wide", initial_sidebar_state="collapsed")

# âââ CUSTOM CSS â Modern Mobile-First Design âââââââââââââââââââââââââââââââââ
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    /* ââ Reset & Base ââ */
    #MainMenu, footer, header {visibility: hidden;}
    html, body, [data-testid="stAppViewContainer"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        background: #f5f6f8;
    }
    .block-container {
        padding: 1.5rem 1rem 5rem 1rem;
        max-width: 800px;
    }
    [data-testid="stSidebar"] {
        min-width: 220px; max-width: 280px;
        background: #0f1d2e;
    }
    [data-testid="stSidebar"] * {color: #c8d6e5 !important;}
    [data-testid="stSidebar"] label[data-testid="stWidgetLabel"] {display: none;}
    [data-testid="stSidebar"] .stRadio > div {gap: 2px;}
    [data-testid="stSidebar"] .stRadio > div > label {
        padding: 10px 16px; border-radius: 8px; font-size: 14px; font-weight: 500;
        transition: background 0.15s;
    }
    [data-testid="stSidebar"] .stRadio > div > label:hover {background: rgba(255,255,255,0.08);}
    [data-testid="stSidebar"] .stRadio > div > label[data-checked="true"] {
        background: rgba(255,255,255,0.12); color: white !important; font-weight: 600;
    }

    /* ââ Typography ââ */
    h1, h2, h3, h4 {font-family: 'Inter', sans-serif; letter-spacing: -0.02em;}
    .page-title {
        font-size: 22px; font-weight: 700; color: #0f1d2e;
        margin: 0 0 4px 0; letter-spacing: -0.03em;
    }
    .page-subtitle {font-size: 13px; color: #8494a7; margin: 0 0 20px 0; font-weight: 400;}

    /* ââ Cards ââ */
    .sf-card {
        background: white; border-radius: 14px; padding: 18px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06), 0 1px 2px rgba(0,0,0,0.04);
        margin-bottom: 12px; border: 1px solid rgba(0,0,0,0.04);
        transition: box-shadow 0.15s;
    }
    .sf-card:hover {box-shadow: 0 4px 12px rgba(0,0,0,0.08), 0 1px 2px rgba(0,0,0,0.04);}

    /* ââ Metric Cards ââ */
    .metric-card {
        background: white; border-radius: 14px; padding: 16px 12px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06); text-align: center;
        border: 1px solid rgba(0,0,0,0.04); border-left: none;
        position: relative; overflow: hidden;
    }
    .metric-card::before {
        content: ''; position: absolute; left: 0; top: 0; bottom: 0;
        width: 4px; background: #1a56db; border-radius: 4px 0 0 4px;
    }
    .metric-value {font-size: 26px; font-weight: 700; color: #0f1d2e; line-height: 1.1;}
    .metric-label {font-size: 11px; color: #8494a7; margin-top: 4px; font-weight: 500; text-transform: uppercase; letter-spacing: 0.05em;}

    /* ââ Badges ââ */
    .sf-badge {
        display: inline-flex; align-items: center; padding: 4px 10px;
        border-radius: 20px; font-size: 12px; font-weight: 600;
        letter-spacing: -0.01em;
    }
    .badge-drywall {background: #dcfce7; color: #166534;}
    .badge-paint {background: #dbeafe; color: #1e40af;}
    .badge-qc {background: #fef3c7; color: #92400e;}
    .badge-ho {background: #e0e7ff; color: #3730a3;}
    .badge-waiting {background: #fef9c3; color: #854d0e;}
    .badge-complete {background: #d1fae5; color: #065f46;}
    .badge-pending {background: #fef3c7; color: #92400e;}
    .badge-confirmed {background: #dcfce7; color: #166534;}
    .badge-denied {background: #fee2e2; color: #991b1b;}
    .badge-discount {background: #ffe4e6; color: #9f1239;}
    .badge-notstarted {background: #f3f4f6; color: #6b7280;}

    /* ââ Community Cards ââ */
    .comm-card {
        background: white; border-radius: 14px; padding: 16px 18px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06); margin-bottom: 10px;
        border: 1px solid rgba(0,0,0,0.04);
    }
    .comm-header {display: flex; justify-content: space-between; align-items: center; margin-bottom: 10px;}
    .comm-name {font-size: 15px; font-weight: 600; color: #0f1d2e;}
    .comm-builder {font-size: 12px; color: #8494a7; font-weight: 500;}
    .comm-count {font-size: 12px; color: #1a56db; font-weight: 600; background: #eff6ff; padding: 3px 10px; border-radius: 20px;}
    .comm-stats {display: flex; gap: 6px; flex-wrap: wrap;}
    .comm-stat {
        text-align: center; padding: 6px 10px; border-radius: 10px;
        min-width: 48px; font-size: 12px;
    }
    .comm-stat-val {font-weight: 700; font-size: 15px; line-height: 1.2;}
    .comm-stat-label {font-size: 9px; font-weight: 500; text-transform: uppercase; letter-spacing: 0.04em; opacity: 0.8;}

    /* ââ Forms ââ */
    div[data-testid="stForm"] {
        background: white; border-radius: 14px; padding: 20px;
        border: 1px solid rgba(0,0,0,0.06);
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }
    .stTextInput > div > div > input,
    .stSelectbox > div > div,
    .stTextArea > div > div > textarea {
        border-radius: 10px !important; border: 1.5px solid #e2e8f0 !important;
        font-size: 14px !important; font-family: 'Inter', sans-serif !important;
    }
    .stTextInput > div > div > input:focus,
    .stTextArea > div > div > textarea:focus {
        border-color: #1a56db !important; box-shadow: 0 0 0 3px rgba(26,86,219,0.1) !important;
    }

    /* ââ Buttons ââ */
    .stButton > button {
        border-radius: 10px; font-family: 'Inter', sans-serif;
        font-weight: 600; font-size: 14px; padding: 8px 20px;
        transition: all 0.15s; letter-spacing: -0.01em;
    }
    .stButton > button[kind="primary"] {
        background: #1a56db; border: none; color: white;
    }
    .stButton > button[kind="primary"]:hover {
        background: #1e40af; box-shadow: 0 4px 12px rgba(26,86,219,0.3);
    }

    /* ââ Alerts ââ */
    .sf-alert {
        padding: 14px 16px; border-radius: 12px; margin: 12px 0;
        font-size: 13px; font-weight: 500; display: flex; align-items: center; gap: 10px;
    }
    .sf-alert-warn {background: #fef9c3; color: #854d0e; border: 1px solid #fde68a;}
    .sf-alert-success {background: #dcfce7; color: #166534; border: 1px solid #bbf7d0;}
    .sf-alert-info {background: #eff6ff; color: #1e40af; border: 1px solid #bfdbfe;}

    /* ââ Success Box ââ */
    .success-box {
        background: #dcfce7; padding: 14px 16px; border-radius: 12px;
        margin: 10px 0; border: 1px solid #bbf7d0; color: #166534;
        font-size: 13px; font-weight: 500;
    }

    /* ââ Expanders ââ */
    .streamlit-expanderHeader {
        font-size: 14px; font-weight: 600; font-family: 'Inter', sans-serif;
        border-radius: 12px;
    }

    /* ââ EPO Row ââ */
    .epo-row {
        background: white; border-radius: 12px; padding: 14px 16px;
        margin-bottom: 8px; border: 1px solid rgba(0,0,0,0.06);
        box-shadow: 0 1px 2px rgba(0,0,0,0.04);
    }
    .epo-header {display: flex; justify-content: space-between; align-items: center; margin-bottom: 4px;}
    .epo-lot {font-size: 14px; font-weight: 600; color: #0f1d2e;}
    .epo-meta {font-size: 12px; color: #8494a7; font-weight: 400;}

    /* ââ Lot Row ââ */
    .lot-row {
        display: flex; justify-content: space-between; align-items: center;
        padding: 12px 16px; background: white; border-radius: 12px;
        margin-bottom: 6px; border: 1px solid rgba(0,0,0,0.04);
        box-shadow: 0 1px 2px rgba(0,0,0,0.03);
    }
    .lot-num {font-size: 14px; font-weight: 600; color: #0f1d2e; min-width: 50px;}
    .lot-notes {font-size: 11px; color: #8494a7; margin-top: 2px;}

    /* ââ Dividers ââ */
    hr {border: none; border-top: 1px solid #e5e7eb; margin: 20px 0;}

    /* ââ Footer ââ */
    .sf-footer {
        text-align: center; font-size: 11px; color: #8494a7;
        padding: 16px 0; margin-top: 20px; font-weight: 500;
    }
</style>
""", unsafe_allow_html=True)

# âââ CONSTANTS ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
DATA_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(DATA_DIR, "data.json")
CONFIG_FILE = os.path.join(DATA_DIR, "config.json")
TRACKER_EMAIL = "stancil.field.tracker@gmail.com"
DW_ORDERS_EMAIL = "dw_orders@stancilservices.com"
DW_CC_EMAILS = "gabriel.jordao@stancilservices.com, alexis.guzman@stancilservices.com"

ALL_STAGES = [
    "Hang","Scrap","Tape","Bed","Skim","Sand",
    "Prime","Waiting for Prime","1st Point Up","1st Paint",
    "Waiting for Trim","Waiting for Final","Final Point Up","Final Paint",
    "Waiting for QC","QC Point Up","QC Paint",
    "Waiting for Homeowners","Homeowners Point Up","Homeowners Paint","Complete",
]
DRYWALL_STAGES = ["Hang","Scrap","Tape","Bed","Skim","Sand"]
PAINT_STAGES = ["Prime","1st Point Up","1st Paint","Final Point Up","Final Paint"]
QC_STAGES = ["QC Point Up","QC Paint"]
HO_STAGES = ["Homeowners Point Up","Homeowners Paint"]
WAITING_STAGES = ["Waiting for Prime","Waiting for Trim","Waiting for Final","Waiting for QC","Waiting for Homeowners"]

# Lot code prefixes and formatting per community
LOT_CODES = {
    "Olmsted":            {"prefix": "PGO0",  "pad": 3, "suffix": ""},
    "Odell Park":         {"prefix": "DRBOP", "pad": 4, "suffix": ""},
    "Mallard Park":       {"prefix": "PGMP0", "pad": 0, "suffix": ""},
    "Anderson Townhomes": {"prefix": "DRHAN", "pad": 0, "suffix": ""},
    "Plot":               {"prefix": "RCPLT", "pad": 0, "suffix": ""},
    "Sugar Creek":        {"prefix": "RCCSC", "pad": 0, "suffix": ""},
    "Galloway":           {"prefix": "PGGW",  "pad": 0, "suffix": "", "needs_building": True},
}

def get_lot_code(community, lot_num, building_num=""):
    cfg = LOT_CODES.get(community)
    if not cfg:
        return lot_num
    prefix = cfg["prefix"]
    pad = cfg.get("pad", 0)
    suffix = cfg.get("suffix", "")
    if cfg.get("needs_building"):
        return f"{prefix}{lot_num}{building_num}"
    if pad > 0:
        try:
            lot_padded = str(int(lot_num)).zfill(pad)
        except ValueError:
            lot_padded = lot_num
    else:
        lot_padded = lot_num
    return f"{prefix}{lot_padded}{suffix}"

PAINT_SUBS = ["GP Painting Services","Jorge Gomez","Christian Painting","Carlos Gabriel","Juan Ulloa"]
POINTUP_SUBS = {
    "Odell Park":"Luis A. Lopez","Mallard Park":"Luis A. Lopez",
    "Galloway":"Luis A. Lopez","Olmsted":"Luis A. Lopez",
    "Sugar Creek":"Edwin","Plot":"Edwin","Anderson Townhomes":"Luis A. Lopez",
}

# âââ DEFAULT DATA âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
DEFAULT_COMMUNITIES = {
    "Odell Park": {"builder":"DRB","lots":{
        "1":{"stage":"Waiting for QC","notes":"Repairs need to be sanded and painted"},
        "2":{"stage":"Waiting for QC","notes":""},"3":{"stage":"Waiting for QC","notes":""},
        "4":{"stage":"Waiting for QC","notes":""},"5":{"stage":"Waiting for QC","notes":""},
        "42":{"stage":"","notes":""},
    },"subs":{"Hang":"America Drywall","Scrap":"America Drywall","Tape":"America Drywall","Bed":"America Drywall","Skim":"America Drywall","Sand":"America Drywall"},
      "durations":{"Hang":1,"Scrap":1,"Tape":2,"Bed":2,"Skim":2,"Sand":1}},
    "Mallard Park": {"builder":"Pulte","lots":{
        "13":{"stage":"Waiting for Homeowners","notes":""},"14":{"stage":"Waiting for Homeowners","notes":""},
        "15":{"stage":"Waiting for Homeowners","notes":""},"16":{"stage":"Waiting for Homeowners","notes":""},
        "17":{"stage":"Waiting for Homeowners","notes":""},"18":{"stage":"Waiting for Homeowners","notes":""},
        "19":{"stage":"Waiting for Homeowners","notes":""},"20":{"stage":"Waiting for Homeowners","notes":""},
        "58":{"stage":"Waiting for Final","notes":""},"59":{"stage":"Prime","notes":""},
        "60":{"stage":"Prime","notes":""},"61":{"stage":"Prime","notes":""},
        "62":{"stage":"Waiting for Final","notes":""},
    },"subs":{t:"America Drywall" for t in DRYWALL_STAGES},"durations":{"Hang":1,"Scrap":1,"Tape":2,"Bed":2,"Skim":2,"Sand":1}},
    "Galloway": {"builder":"Pulte","lots":{
        "21":{"stage":"","notes":""},"22":{"stage":"","notes":""},
        "23":{"stage":"","notes":""},"24":{"stage":"","notes":""},
        "25":{"stage":"Waiting for Final","notes":""},"26":{"stage":"Waiting for Final","notes":""},
        "27":{"stage":"Waiting for Final","notes":""},"28":{"stage":"Waiting for Final","notes":""},
    },"subs":{t:"America Drywall" for t in DRYWALL_STAGES},"durations":{"Hang":1,"Scrap":1,"Tape":2,"Bed":2,"Skim":2,"Sand":1}},
    "Olmsted": {"builder":"Pote","lots":{
        "276":{"stage":"Waiting for Final","notes":""},"277":{"stage":"1st Paint","notes":""},
        "278":{"stage":"Prime","notes":""},"279":{"stage":"Prime","notes":""},
        "280":{"stage":"Hang","notes":""},
    },"subs":{"Hang":"Ricardo","Scrap":"Scrap Brothers","Tape":"Juan Trejo","Bed":"Juan Trejo","Skim":"Juan Trejo","Sand":"Juan Trejo"},
      "durations":{"Hang":1,"Scrap":1,"Tape":2,"Bed":2,"Skim":2,"Sand":1}},
    "Sugar Creek": {"builder":"Red Cedar","lots":{
        **{str(i):{"stage":"Waiting for Final","notes":""} for i in range(1,34)},
        "12":{"stage":"Waiting for Final","notes":"Need to ask for an EPO for accent colors"},
        "34":{"stage":"Prime","notes":""},"35":{"stage":"Prime","notes":""},"36":{"stage":"Prime","notes":""},
        "37":{"stage":"Waiting for Trim","notes":""},"38":{"stage":"Prime","notes":""},
        "39":{"stage":"Waiting for Trim","notes":""},"40":{"stage":"Prime","notes":""},
    },"subs":{t:"America Drywall" for t in DRYWALL_STAGES},"durations":{"Hang":1,"Scrap":1,"Tape":2,"Bed":2,"Skim":2,"Sand":1}},
    "Plot": {"builder":"Red Cedar","lots":{
        "1A":{"stage":"Waiting for QC","notes":"Need to ask for EPO for accent colors"},
        "1B":{"stage":"Waiting for QC","notes":""},"1C":{"stage":"Waiting for QC","notes":""},"1D":{"stage":"Waiting for QC","notes":""},
        "2A":{"stage":"Waiting for QC","notes":""},"2B":{"stage":"Waiting for QC","notes":""},
        "2C":{"stage":"Waiting for QC","notes":""},"2D":{"stage":"Waiting for QC","notes":""},
        "3A":{"stage":"Final Paint","notes":""},"3B":{"stage":"Final Paint","notes":""},
        "3C":{"stage":"Final Paint","notes":""},"3D":{"stage":"Final Paint","notes":""},
        "4A":{"stage":"Final Paint","notes":""},"4B":{"stage":"Final Paint","notes":""},
        "4C":{"stage":"Final Paint","notes":""},"4D":{"stage":"Final Paint","notes":""},
        "5A":{"stage":"","notes":""},"5B":{"stage":"","notes":""},"5C":{"stage":"","notes":""},"5D":{"stage":"","notes":""},
        "6A":{"stage":"Waiting for Trim","notes":""},"6B":{"stage":"Waiting for Trim","notes":""},
        "6C":{"stage":"Waiting for Trim","notes":""},"6D":{"stage":"Waiting for Trim","notes":""},
        "7A":{"stage":"Sand","notes":""},"7B":{"stage":"Sand","notes":""},
        "7C":{"stage":"Sand","notes":""},"7D":{"stage":"Sand","notes":""},
        "8A":{"stage":"Waiting for Final","notes":""},"8B":{"stage":"Waiting for Final","notes":""},
        "8C":{"stage":"Waiting for Final","notes":""},"8D":{"stage":"Waiting for Final","notes":""},
    },"subs":{t:"America Drywall" for t in DRYWALL_STAGES},"durations":{"Hang":1,"Scrap":1,"Tape":2,"Bed":2,"Skim":2,"Sand":1}},
    "Anderson Townhomes": {"builder":"DR Horton","lots":{
        "15":{"stage":"","notes":""},"16":{"stage":"","notes":""},
        "17":{"stage":"","notes":""},"18":{"stage":"","notes":""},"19":{"stage":"","notes":""},
    },"subs":{t:"America Drywall" for t in DRYWALL_STAGES},"durations":{"Hang":1,"Scrap":1,"Tape":2,"Bed":2,"Skim":2,"Sand":1}},
}

# âââ DATA / CONFIG ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE) as f: return json.load(f)
    return {"communities":copy.deepcopy(DEFAULT_COMMUNITIES),"epo_log":[],"notes":[],"schedules":[]}

def save_data(data):
    with open(DATA_FILE,"w") as f: json.dump(data,f,default=str,indent=2)

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE) as f: return json.load(f)
    return {"gmail_app_password":""}

def save_config(cfg):
    with open(CONFIG_FILE,"w") as f: json.dump(cfg,f,indent=2)

if "data" not in st.session_state:
    st.session_state.data = load_data()
if "config" not in st.session_state:
    st.session_state.config = load_config()

def get_data(): return st.session_state.data
def persist(): save_data(st.session_state.data)

# âââ EMAIL ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
def send_email(to, subject, body, cc=None):
    cfg = st.session_state.config
    app_pw = cfg.get("gmail_app_password","")
    if not app_pw:
        return False, "Gmail app password not configured. Go to Settings to set it up."
    try:
        msg = MIMEMultipart()
        msg["From"] = TRACKER_EMAIL
        msg["To"] = to
        msg["Subject"] = subject
        if cc:
            msg["Cc"] = cc if isinstance(cc, str) else ", ".join(cc)
        msg.attach(MIMEText(body, "plain"))
        # Build full recipient list
        recipients = [to]
        if cc:
            cc_list = [c.strip() for c in (cc.split(",") if isinstance(cc, str) else cc)]
            recipients.extend(cc_list)
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(TRACKER_EMAIL, app_pw)
            server.sendmail(TRACKER_EMAIL, recipients, msg.as_string())
        return True, "Email sent!"
    except Exception as e:
        return False, f"Failed to send: {str(e)}"

# âââ HELPERS ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
def count_by_category(lots):
    c = {"drywall":0,"paint":0,"qc":0,"ho":0,"waiting":0,"complete":0,"not_started":0}
    for ld in lots.values():
        s = ld.get("stage","")
        if not s: c["not_started"]+=1
        elif s in DRYWALL_STAGES: c["drywall"]+=1
        elif s in PAINT_STAGES: c["paint"]+=1
        elif s in QC_STAGES: c["qc"]+=1
        elif s in HO_STAGES: c["ho"]+=1
        elif s in WAITING_STAGES: c["waiting"]+=1
        elif s=="Complete": c["complete"]+=1
    return c

def _epo_days_open(epo):
    """Calculate number of days since an EPO was submitted."""
    date_str = epo.get("date", epo.get("sent", ""))
    try:
        sent_date = datetime.datetime.strptime(date_str.split(" ")[0], "%m/%d/%Y").date()
        return (datetime.date.today() - sent_date).days
    except Exception:
        return 0

# âââ BUILDER CONTACT MAP (To â community mapping) ââââââââââââââââââââââââââââ
BUILDER_CONTACTS = {
    "angel.serrano@pulte.com": {"name": "Angel Serrano", "builder": "Pulte", "communities": ["Mallard Park", "Galloway"]},
    "gus.schnitker@pulte.com": {"name": "Gus Schnitker", "builder": "Pulte", "communities": ["Mallard Park", "Galloway"]},
    # Add more contacts as needed in Settings
}

def _parse_epo_subject(subject):
    """Extract lot number and community from EPO email subject."""
    subject_lower = subject.lower().strip()
    lot_num = ""
    community = ""
    # Extract lot number: "lot 20", "lot20", "lot #20", "lot# 20"
    lot_match = re.search(r'lot\s*#?\s*(\w+)', subject_lower)
    if lot_match:
        lot_num = lot_match.group(1).upper() if any(c.isalpha() for c in lot_match.group(1)) else lot_match.group(1)
    # Extract community name by checking known communities
    for comm_name in LOT_CODES.keys():
        if comm_name.lower() in subject_lower:
            community = comm_name
            break
    return lot_num, community

def _parse_epo_body(body_text):
    """Extract amount and description from EPO email body."""
    amount = ""
    description = ""
    lines = body_text.strip().splitlines()
    # Get the actual message (before signature block)
    msg_lines = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        # Stop at signature indicators
        if any(sig in stripped.lower() for sig in ["field manager", "residential painting", "stancil", "www.stancil"]):
            break
        if stripped.startswith("--") or stripped.startswith("___"):
            break
        msg_lines.append(stripped)
    msg_text = " ".join(msg_lines)
    # Extract amount: "$350", "of 350", "of $350", "$1,200"
    amt_match = re.search(r'\$\s?([\d,]+(?:\.\d{2})?)', msg_text)
    if amt_match:
        amount = amt_match.group(1).replace(",", "")
    else:
        amt_match2 = re.search(r'(?:epo\s+(?:of|for)\s+)([\d,]+(?:\.\d{2})?)', msg_text, re.IGNORECASE)
        if amt_match2:
            amount = amt_match2.group(1).replace(",", "")
    # Description is the full message text (cleaned up)
    description = msg_text
    # Remove "epo of 350" prefix pattern to get just the description
    desc_clean = re.sub(r'^.*?epo\s+(?:of|for)\s+\$?[\d,]+\s*(?:to\s+)?', '', msg_text, flags=re.IGNORECASE).strip()
    if desc_clean:
        description = desc_clean
    return amount, description

def _get_email_text(msg):
    """Extract plain text body from an email message."""
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            if ctype == "text/plain":
                try:
                    return part.get_payload(decode=True).decode("utf-8", errors="replace")
                except:
                    return part.get_payload(decode=True).decode("latin-1", errors="replace")
    else:
        try:
            return msg.get_payload(decode=True).decode("utf-8", errors="replace")
        except:
            return msg.get_payload(decode=True).decode("latin-1", errors="replace")
    return ""

def _decode_header_value(val):
    """Decode an email header value."""
    if val is None:
        return ""
    decoded = decode_header(val)
    parts = []
    for content, charset in decoded:
        if isinstance(content, bytes):
            parts.append(content.decode(charset or "utf-8", errors="replace"))
        else:
            parts.append(content)
    return " ".join(parts)

def fetch_epo_emails(data):
    """Read Gmail inbox via IMAP and parse EPO emails."""
    cfg = load_config()
    app_pw = cfg.get("gmail_app_password", "")
    if not app_pw:
        return [], "No Gmail app password configured. Go to Settings first."
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(TRACKER_EMAIL, app_pw)
        mail.select("inbox")
        # Search for emails with "epo" in subject
        status, msg_ids = mail.search(None, '(SUBJECT "epo")')
        if status != "OK" or not msg_ids[0]:
            mail.logout()
            return [], "No EPO emails found in inbox."
        # Get already-processed message IDs
        processed_ids = set(e.get("email_id", "") for e in data.get("epo_log", []) if e.get("email_id"))
        new_epos = []
        id_list = msg_ids[0].split()
        # Process most recent first, limit to last 100
        for mid in reversed(id_list[-100:]):
            mid_str = mid.decode()
            # Fetch message
            status, msg_data = mail.fetch(mid, "(RFC822 BODY[HEADER.FIELDS (MESSAGE-ID)])")
            if status != "OK":
                continue
            raw_email = msg_data[0][1]
            msg = email_lib.message_from_bytes(raw_email)
            # Get message ID for deduplication
            message_id = msg.get("Message-ID", mid_str)
            if message_id in processed_ids:
                continue
            subject = _decode_header_value(msg.get("Subject", ""))
            if "epo" not in subject.lower():
                continue
            from_addr = msg.get("From", "")
            to_raw = msg.get("To", "")
            date_str = msg.get("Date", "")
            body = _get_email_text(msg)
            # Parse fields
            lot_num, community = _parse_epo_subject(subject)
            amount, description = _parse_epo_body(body)
            # Extract builder contact from To field
            to_match = re.search(r'([^<\s]+@[^>\s]+)', to_raw)
            builder_contact_email = to_match.group(1).lower() if to_match else ""
            # Try to get builder contact name
            name_match = re.search(r'^([^<]+)', to_raw)
            builder_contact_name = name_match.group(1).strip().strip('"') if name_match else ""
            # Look up builder from contact map or community data
            builder = ""
            contact_info = BUILDER_CONTACTS.get(builder_contact_email, {})
            if contact_info:
                builder = contact_info.get("builder", "")
                if not community and contact_info.get("communities"):
                    community = contact_info["communities"][0]
                if not builder_contact_name:
                    builder_contact_name = contact_info.get("name", "")
            # If we still don't have a community, try matching from known data
            if not community:
                for comm_name, comm_data in data.get("communities", {}).items():
                    if builder and comm_data.get("builder", "").lower() == builder.lower():
                        community = comm_name
                        break
            # Parse email date
            try:
                from email.utils import parsedate_to_datetime
                dt = parsedate_to_datetime(date_str)
                parsed_date = dt.strftime("%m/%d/%Y")
                parsed_timestamp = dt.strftime("%m/%d/%Y %H:%M")
            except:
                parsed_date = datetime.datetime.now().strftime("%m/%d/%Y")
                parsed_timestamp = datetime.datetime.now().strftime("%m/%d/%Y %H:%M")
            lot_code = get_lot_code(community, lot_num) if community and lot_num else ""
            epo_entry = {
                "email_id": message_id,
                "field_manager": "GABRIEL JORDAO",
                "builder_contact": builder_contact_name,
                "builder_contact_email": builder_contact_email,
                "date": parsed_date,
                "builder": builder,
                "neighborhood": community,
                "lot": lot_num,
                "short_name": lot_code,
                "description": description,
                "amount": amount,
                "status": "Pending",
                "confirmation_num": "",
                "followup_count": 0,
                "sent_timestamp": parsed_timestamp,
                "needs_review": not all([lot_num, community, amount]),
                "original_subject": subject,
            }
            new_epos.append(epo_entry)
        mail.logout()
        return new_epos, f"Found {len(new_epos)} new EPO email(s)."
    except Exception as e:
        return [], f"IMAP error: {str(e)}"

def generate_schedule(comm_name, lot_num, start_date, data):
    comm = data["communities"].get(comm_name,{})
    durations = comm.get("durations",{"Hang":1,"Scrap":1,"Tape":2,"Bed":2,"Skim":2,"Sand":1})
    subs = comm.get("subs",{})
    schedule, cur = [], start_date
    for task in DRYWALL_STAGES:
        days, added = durations.get(task,1), 0
        while added < days:
            cur += datetime.timedelta(days=1)
            if task=="Scrap" and cur.weekday()>=5: continue
            if task!="Scrap" and cur.weekday()==6: continue
            added += 1
        schedule.append({"task":task,"sub":subs.get(task,"TBD"),"date":cur.strftime("%m/%d/%Y"),"date_obj":cur})
    return schedule

# âââ SHARED REPORT VIEW ââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# If ?view=report is in the URL, show read-only dashboard for area manager
query_params = st.query_params
if query_params.get("view") == "report":
    data = get_data()
    comms = data["communities"]
    st.markdown('<div style="text-align:center;padding:20px 0 10px;">'
                '<div style="font-size:22px;font-weight:700;color:#0f1d2e;letter-spacing:-0.03em;">Stancil Field Manager</div>'
                '<div style="font-size:13px;color:#8494a7;margin-top:2px;">Job Report</div></div>', unsafe_allow_html=True)
    st.markdown(f'<div style="text-align:center;font-size:13px;color:#8494a7;margin-bottom:20px;">Week of {datetime.date.today().strftime("%B %d, %Y")} Â· Gabriel Jordao</div>', unsafe_allow_html=True)

    total_lots = sum(len(c["lots"]) for c in comms.values())
    total_active = sum(1 for c in comms.values() for l in c["lots"].values() if l.get("stage") and l["stage"]!="Complete")
    total_waiting = sum(1 for c in comms.values() for l in c["lots"].values() if l.get("stage","") in WAITING_STAGES)
    total_complete = sum(1 for c in comms.values() for l in c["lots"].values() if l.get("stage")=="Complete")

    c1,c2,c3,c4 = st.columns(4)
    c1.markdown(f'<div class="metric-card"><div class="metric-value">{total_lots}</div><div class="metric-label">Total Lots</div></div>',unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-card"><div class="metric-value">{total_active}</div><div class="metric-label">Active</div></div>',unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#d97706">{total_waiting}</div><div class="metric-label">Waiting</div></div>',unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#059669">{total_complete}</div><div class="metric-label">Complete</div></div>',unsafe_allow_html=True)

    st.markdown("---")

    for comm_name, comm_data in comms.items():
        counts = count_by_category(comm_data["lots"])
        total = len(comm_data["lots"])
        active = total - counts["complete"] - counts["not_started"]
        stats_html = ""
        for label, key, bg, fg in [("DW","drywall","#dcfce7","#166534"),("Paint","paint","#dbeafe","#1e40af"),
                                    ("QC","qc","#fef3c7","#92400e"),("HO","ho","#e0e7ff","#3730a3"),
                                    ("Wait","waiting","#fef9c3","#854d0e"),("Done","complete","#d1fae5","#065f46")]:
            stats_html += f'<div class="comm-stat" style="background:{bg};color:{fg};"><div class="comm-stat-val">{counts[key]}</div><div class="comm-stat-label">{label}</div></div>'
        st.markdown(f'''<div class="comm-card">
            <div class="comm-header">
                <div><span class="comm-name">{comm_name}</span><br><span class="comm-builder">{comm_data["builder"]}</span></div>
                <span class="comm-count">{active} / {total}</span>
            </div>
            <div class="comm-stats">{stats_html}</div>
        </div>''', unsafe_allow_html=True)

        # Show every lot with stage and notes
        lots_html = ""
        for lot_num, lot_data in comm_data["lots"].items():
            stage = lot_data.get("stage", "")
            notes = lot_data.get("notes", "")
            b = badge_class = ""
            if not stage:
                b_bg, b_fg, b_text = "#f1f5f9", "#94a3b8", "Not Started"
            elif stage in DRYWALL_STAGES:
                b_bg, b_fg, b_text = "#dcfce7", "#166534", stage
            elif stage in PAINT_STAGES:
                b_bg, b_fg, b_text = "#dbeafe", "#1e40af", stage
            elif stage in QC_STAGES:
                b_bg, b_fg, b_text = "#fef3c7", "#92400e", stage
            elif stage in HO_STAGES:
                b_bg, b_fg, b_text = "#e0e7ff", "#3730a3", stage
            elif stage in WAITING_STAGES:
                b_bg, b_fg, b_text = "#fef9c3", "#854d0e", stage
            elif stage == "Complete":
                b_bg, b_fg, b_text = "#d1fae5", "#065f46", stage
            else:
                b_bg, b_fg, b_text = "#f1f5f9", "#64748b", stage
            notes_html = f'<div style="font-size:11px;color:#8494a7;margin-top:2px;">{notes}</div>' if notes else ""
            lots_html += f'''<div class="lot-row">
                <div><span class="lot-num">Lot {lot_num}</span>{notes_html}</div>
                <span class="sf-badge" style="background:{b_bg};color:{b_fg};">{b_text}</span>
            </div>'''
        st.markdown(lots_html, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

    st.markdown('<div class="sf-footer">Stancil Field Manager v2.0 Â· Shared Report</div>', unsafe_allow_html=True)
    st.stop()

# âââ NAV ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# TOP NAVIGATION BAR
st.markdown('<div style="font-size:20px;font-weight:700;color:#0f1d2e;letter-spacing:-0.03em;margin-bottom:2px;">Stancil <span style="font-weight:400;color:#8494a7;font-size:14px;">Field Manager</span></div>', unsafe_allow_html=True)

nav_options = ["Dashboard","Job Tracker","Order","EPO Tracker","Field Notes","Settings"]

if "current_page" not in st.session_state:
    st.session_state.current_page = "Dashboard"

cols = st.columns(len(nav_options))
for i, opt in enumerate(nav_options):
    with cols[i]:
        if st.button(opt, key=f"nav_{opt}", use_container_width=True,
                     type="primary" if st.session_state.current_page == opt else "secondary"):
            st.session_state.current_page = opt
            st.rerun()

page = st.session_state.current_page
st.markdown("---")

# âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# DASHBOARD
# âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
if page == "Dashboard":
    st.markdown('<div class="page-title">Dashboard</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="page-subtitle">Week of {datetime.date.today().strftime("%B %d, %Y")}</div>', unsafe_allow_html=True)
    data = get_data()
    comms = data["communities"]
    total_lots = sum(len(c["lots"]) for c in comms.values())
    total_active = sum(1 for c in comms.values() for l in c["lots"].values() if l.get("stage") and l["stage"]!="Complete")
    total_waiting = sum(1 for c in comms.values() for l in c["lots"].values() if l.get("stage","") in WAITING_STAGES)
    total_complete = sum(1 for c in comms.values() for l in c["lots"].values() if l.get("stage")=="Complete")

    c1,c2,c3,c4 = st.columns(4)
    c1.markdown(f'<div class="metric-card"><div class="metric-value">{total_lots}</div><div class="metric-label">Total Lots</div></div>',unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-card"><div class="metric-value">{total_active}</div><div class="metric-label">Active</div></div>',unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#d97706">{total_waiting}</div><div class="metric-label">Waiting</div></div>',unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#059669">{total_complete}</div><div class="metric-label">Complete</div></div>',unsafe_allow_html=True)

    # EPO alerts at top
    epo_log = data.get("epo_log", [])
    epo_pending = [e for e in epo_log if e.get("status", "Pending") == "Pending"]
    epo_overdue = [e for e in epo_pending if _epo_days_open(e) >= 4]
    epo_confirmed = len([e for e in epo_log if e.get("status") == "Confirmed"])
    epo_total = len(epo_log)
    capture_rate = f"{(epo_confirmed/epo_total*100):.0f}%" if epo_total > 0 else "â"

    if epo_overdue:
        st.markdown(f'<div class="sf-alert sf-alert-warn">â &nbsp;&nbsp;<strong>EPO Follow-Up Needed ({len(epo_overdue)})</strong></div>', unsafe_allow_html=True)
        for e in epo_overdue:
            days = _epo_days_open(e)
            st.markdown(f'<div class="epo-row"><div class="epo-header"><span class="epo-lot">Lot {e["lot"]} â {e.get("neighborhood","")}</span>'
                        f'<span class="sf-badge badge-pending">{days}d open</span></div>'
                        f'<div class="epo-meta">${e.get("amount","")} Â· Follow up needed</div></div>', unsafe_allow_html=True)

    st.markdown("---")

    # Communities
    for comm_name, comm_data in comms.items():
        counts = count_by_category(comm_data["lots"])
        total = len(comm_data["lots"])
        active = total - counts["complete"] - counts["not_started"]
        stats_html = ""
        for label, key, bg, fg in [("DW",  "drywall", "#dcfce7","#166534"),
                                    ("Paint","paint",   "#dbeafe","#1e40af"),
                                    ("QC",   "qc",      "#fef3c7","#92400e"),
                                    ("HO",   "ho",      "#e0e7ff","#3730a3"),
                                    ("Wait", "waiting", "#fef9c3","#854d0e"),
                                    ("Done", "complete","#d1fae5","#065f46")]:
            stats_html += f'<div class="comm-stat" style="background:{bg};color:{fg};"><div class="comm-stat-val">{counts[key]}</div><div class="comm-stat-label">{label}</div></div>'
        st.markdown(f'''<div class="comm-card">
            <div class="comm-header">
                <div><span class="comm-name">{comm_name}</span><br><span class="comm-builder">{comm_data["builder"]}</span></div>
                <span class="comm-count">{active} / {total}</span>
            </div>
            <div class="comm-stats">{stats_html}</div>
        </div>''', unsafe_allow_html=True)

    # EPO summary
    if epo_total > 0:
        st.markdown("---")
        st.markdown(f'<div class="sf-card"><div style="display:flex;justify-content:space-between;align-items:center;">'
                    f'<div><span style="font-size:14px;font-weight:600;color:#0f1d2e;">EPO Summary</span><br>'
                    f'<span style="font-size:12px;color:#8494a7;">{len(epo_pending)} pending Â· {epo_confirmed} confirmed</span></div>'
                    f'<span style="font-size:22px;font-weight:700;color:#1a56db;">{capture_rate}</span></div></div>', unsafe_allow_html=True)

    # Share report link
    st.markdown("---")
    st.markdown('<div style="font-size:14px;font-weight:600;color:#0f1d2e;margin-bottom:8px;">Share Report</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:12px;color:#8494a7;margin-bottom:10px;">Share this link with your area manager. '
                'They can see every lot, stage, and notes â read-only, no login needed.</div>', unsafe_allow_html=True)
    # Build the share URL
    share_url = "?view=report"
    st.code(share_url, language=None)
    st.caption("When deployed, this will be your full URL + ?view=report (e.g. https://stancil-field-manager.streamlit.app/?view=report)")

# âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# JOB TRACKER
# âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
elif page == "Job Tracker":
    st.markdown('<div class="page-title">Job Tracker</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-subtitle">Track lots and stages by community</div>', unsafe_allow_html=True)
    data = get_data()
    comms = data["communities"]
    selected_comm = st.selectbox("Select Community", list(comms.keys()))
    comm = comms[selected_comm]
    st.markdown(f"**Builder:** {comm['builder']}")

    with st.expander("Add New Lot"):
        with st.form("add_lot_form",clear_on_submit=True):
            new_lot = st.text_input("Lot Number")
            new_stage = st.selectbox("Stage (optional)",[""] + ALL_STAGES)
            new_notes = st.text_input("Notes (optional)")
            add_btn = st.form_submit_button("Add Lot")
        if add_btn and new_lot:
            comm["lots"][new_lot] = {"stage":new_stage,"notes":new_notes,"updated":datetime.date.today().strftime("%m/%d/%Y")}
            persist(); st.success(f"Lot {new_lot} added!"); st.rerun()

    st.markdown("---")
    for lot_num in list(comm["lots"].keys()):
        lot_data = comm["lots"][lot_num]
        current_stage = lot_data.get("stage","")
        stage_display = current_stage if current_stage else "Not Started"
        note_indicator = " *" if lot_data.get("notes") else ""
        with st.expander(f"Lot {lot_num}  â  {stage_display}{note_indicator}"):
            # Show note history if it exists
            note_history = lot_data.get("note_history", [])
            if note_history:
                st.markdown('<div style="font-size:12px;font-weight:600;color:#0f1d2e;margin-bottom:6px;">Notes History</div>', unsafe_allow_html=True)
                for nh in reversed(note_history[-5:]):
                    st.markdown(f'<div style="font-size:12px;padding:6px 10px;background:#f8fafc;border-radius:8px;margin-bottom:4px;border-left:3px solid #1a56db;">'
                                f'<span style="color:#0f1d2e;">{nh["text"]}</span> '
                                f'<span style="color:#8494a7;font-size:10px;">Â· {nh.get("action","")} Â· {nh.get("timestamp","")}</span></div>', unsafe_allow_html=True)
            elif lot_data.get("notes"):
                st.markdown(f'<div style="font-size:12px;color:#8494a7;margin-bottom:8px;">{lot_data["notes"]}</div>', unsafe_allow_html=True)

            with st.form(f"edit_{selected_comm}_{lot_num}",clear_on_submit=False):
                stage_idx = ALL_STAGES.index(current_stage)+1 if current_stage in ALL_STAGES else 0
                new_stage = st.selectbox("Stage",[""] + ALL_STAGES,index=stage_idx,key=f"s_{selected_comm}_{lot_num}")
                new_notes = st.text_input("Notes",value=lot_data.get("notes",""),key=f"n_{selected_comm}_{lot_num}")
                col1,col2 = st.columns(2)
                save_btn = col1.form_submit_button("Save")
                del_btn = col2.form_submit_button("Remove Lot")
            if save_btn:
                lot_data["stage"] = new_stage
                lot_data["notes"] = new_notes
                lot_data["updated"] = datetime.date.today().strftime("%m/%d/%Y")
                persist(); st.success(f"Lot {lot_num} updated"); st.rerun()
            if del_btn:
                del comm["lots"][lot_num]; persist(); st.success(f"Lot {lot_num} removed"); st.rerun()

    st.markdown("---")
    with st.expander("Add New Community"):
        with st.form("add_comm_form",clear_on_submit=True):
            nc_name = st.text_input("Community Name")
            nc_builder = st.text_input("Builder")
            nc_btn = st.form_submit_button("Add Community")
        if nc_btn and nc_name:
            data["communities"][nc_name] = {"builder":nc_builder,"lots":{},"subs":{t:"TBD" for t in DRYWALL_STAGES},
                "durations":{"Hang":1,"Scrap":1,"Tape":2,"Bed":2,"Skim":2,"Sand":1}}
            persist(); st.success(f"{nc_name} added!"); st.rerun()

# âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# ORDER
# âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
elif page == "Order":
    st.markdown('<div class="page-title">Drywall Order</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-subtitle">Generate and send material order emails</div>', unsafe_allow_html=True)
    data = get_data()
    comms = data["communities"]

    with st.form("order_form"):
        col1, col2 = st.columns(2)
        comm_name = col1.selectbox("Community", list(comms.keys()))
        lot_num = col2.text_input("Lot Number")
        building_num = ""
        if LOT_CODES.get(comm_name,{}).get("needs_building"):
            building_num = st.text_input("Building Number (Galloway)")
        scrap_date = st.date_input("Scrap Date (delivery date)")
        order_btn = st.form_submit_button("Generate Order")

    if order_btn and lot_num:
        lot_code = get_lot_code(comm_name, lot_num, building_num)
        scrap_short = scrap_date.strftime("%m/%d")

        email_subject = f"drywall materials for lot {lot_num} {comm_name.lower()}"
        email_body = f"please have drywall material delivered to {lot_code} on {scrap_short}\nthank you"

        st.markdown(f'''<div class="sf-card">
            <div style="font-size:11px;color:#8494a7;font-weight:500;text-transform:uppercase;letter-spacing:0.05em;margin-bottom:10px;">Order Preview</div>
            <div style="margin-bottom:8px;"><span style="font-size:12px;color:#8494a7;">To:</span> <span style="font-size:13px;font-weight:500;color:#0f1d2e;">{DW_ORDERS_EMAIL}</span></div>
            <div style="margin-bottom:8px;"><span style="font-size:12px;color:#8494a7;">CC:</span> <span style="font-size:13px;font-weight:500;color:#0f1d2e;">{DW_CC_EMAILS}</span></div>
            <div style="margin-bottom:8px;"><span style="font-size:12px;color:#8494a7;">Subject:</span> <span style="font-size:13px;font-weight:500;color:#0f1d2e;">{email_subject}</span></div>
            <div style="background:#f8fafc;border-radius:10px;padding:14px;margin-top:12px;font-size:13px;color:#334155;line-height:1.6;border:1px solid #e2e8f0;">{email_body.replace(chr(10),"<br>")}</div>
        </div>''', unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        if col1.button("Send Order Email", type="primary"):
            success, msg = send_email(DW_ORDERS_EMAIL, email_subject, email_body, cc=DW_CC_EMAILS)
            if success:
                # Log the order
                data.setdefault("orders",[]).append({
                    "community": comm_name, "lot": lot_num, "lot_code": lot_code,
                    "scrap_date": scrap_date.strftime("%m/%d/%Y"),
                    "sent": datetime.datetime.now().strftime("%m/%d/%Y %H:%M"),
                })
                persist()
                st.markdown(f'<div class="success-box">Order sent to {DW_ORDERS_EMAIL}</div>',unsafe_allow_html=True)
            else:
                st.error(msg)
        if col2.button("Copy to Clipboard"):
            st.code(f"To: {DW_ORDERS_EMAIL}\nCC: {DW_CC_EMAILS}\nSubject: {email_subject}\n\n{email_body}", language=None)

    # Recent orders
    orders = data.get("orders", [])
    if orders:
        st.markdown("---")
        st.markdown('<div style="font-size:14px;font-weight:600;color:#0f1d2e;margin-bottom:10px;">Recent Orders</div>', unsafe_allow_html=True)
        for o in reversed(orders[-10:]):
            st.markdown(f'<div class="lot-row"><div><span class="lot-num">Lot {o["lot"]}</span>'
                        f'<div class="lot-notes">{o["community"]} Â· {o["lot_code"]}</div></div>'
                        f'<div style="text-align:right;"><span style="font-size:12px;color:#0f1d2e;font-weight:500;">Deliver {o["scrap_date"]}</span>'
                        f'<div class="lot-notes">Sent {o.get("sent","")}</div></div></div>', unsafe_allow_html=True)

# âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# EPO TRACKER
# âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
elif page == "EPO Tracker":
    st.markdown('<div class="page-title">EPO Tracker</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-subtitle">Track extra paint orders and capture rate</div>', unsafe_allow_html=True)
    data = get_data()
    comms = data["communities"]
    FIELD_MANAGER = "GABRIEL JORDAO"
    epo_log = data.get("epo_log", [])

    # ââ Sync from Gmail inbox ââ
    st.markdown('<div class="sf-alert sf-alert-info">Send EPOs from Outlook as usual â just CC '
                '<strong>stancil.field.tracker@gmail.com</strong>. Hit Sync to pull them in.</div>', unsafe_allow_html=True)

    sync_col1, sync_col2 = st.columns([1,3])
    if sync_col1.button("Sync Inbox", type="primary"):
        with st.spinner("Checking Gmail for new EPO emails..."):
            new_epos, sync_msg = fetch_epo_emails(data)
        if new_epos:
            data.setdefault("epo_log", []).extend(new_epos)
            epo_log = data["epo_log"]
            persist()
            needs_review = len([e for e in new_epos if e.get("needs_review")])
            st.markdown(f'<div class="success-box">Synced {len(new_epos)} new EPO(s)!'
                        f'{"  (" + str(needs_review) + " need review)" if needs_review else ""}</div>',
                        unsafe_allow_html=True)
        else:
            st.info(sync_msg)
    sync_col2.caption(f"Last sync: {data.get('last_epo_sync','Never')}  |  {len(epo_log)} EPOs tracked")
    # Update last sync time
    if "last_epo_sync" not in data or sync_col1.button("x", key="_hidden", disabled=True, label_visibility="hidden") is None:
        pass  # just for layout

    # ââ Capture rate summary ââ
    total_epos = len(epo_log)
    confirmed = len([e for e in epo_log if e.get("status") == "Confirmed"])
    denied = len([e for e in epo_log if e.get("status") == "Denied"])
    pending = len([e for e in epo_log if e.get("status","Pending") == "Pending"])
    discount = len([e for e in epo_log if e.get("status") == "Discount Request"])
    followup_needed = len([e for e in epo_log if e.get("status","Pending") == "Pending" and _epo_days_open(e) >= 4])
    capture_rate = f"{(confirmed/total_epos*100):.0f}%" if total_epos > 0 else "N/A"

    c1,c2,c3,c4 = st.columns(4)
    c1.markdown(f'<div class="metric-card"><div class="metric-value">{total_epos}</div><div class="metric-label">Total EPOs</div></div>',unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#548235">{confirmed}</div><div class="metric-label">Confirmed</div></div>',unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#BF8F00">{pending}</div><div class="metric-label">Pending</div></div>',unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#1F4E79">{capture_rate}</div><div class="metric-label">Capture Rate</div></div>',unsafe_allow_html=True)

    # ââ Follow-up alerts (4+ days) ââ
    if followup_needed > 0:
        st.markdown(f'<div style="background:#FFF2CC;padding:12px;border-radius:8px;border-left:4px solid #BF8F00;margin:10px 0;">'
                    f'<strong>Follow-Up Needed ({followup_needed})</strong></div>',unsafe_allow_html=True)
        for e in epo_log:
            if e.get("status","Pending") == "Pending" and _epo_days_open(e) >= 4:
                days = _epo_days_open(e)
                st.warning(f"**Lot {e.get('lot','?')}** â {e.get('neighborhood','?')} | ${e.get('amount','?')} | {days} days open")

    # ââ Manual EPO entry (fallback) ââ
    st.markdown("---")
    with st.expander("Add EPO Manually (if not sent by email)"):
        with st.form("epo_manual_form", clear_on_submit=True):
            col1,col2 = st.columns(2)
            epo_comm = col1.selectbox("Community", list(comms.keys()))
            epo_lot = col2.text_input("Lot Number")
            epo_contact = st.text_input("Builder Contact Name")
            epo_contact_email = st.text_input("Builder Contact Email")
            epo_amt = st.text_input("Amount ($)")
            epo_desc = st.text_area("Description of work", height=80)
            add_manual = st.form_submit_button("Add EPO")
        if add_manual and epo_lot:
            builder = comms.get(epo_comm,{}).get("builder","")
            lot_code = get_lot_code(epo_comm, epo_lot)
            now_str = datetime.datetime.now().strftime("%m/%d/%Y")
            epo_entry = {
                "field_manager": FIELD_MANAGER,
                "builder_contact": epo_contact,
                "builder_contact_email": epo_contact_email,
                "date": now_str,
                "builder": builder,
                "neighborhood": epo_comm,
                "lot": epo_lot,
                "short_name": lot_code,
                "description": epo_desc,
                "amount": epo_amt,
                "status": "Pending",
                "confirmation_num": "",
                "followup_count": 0,
                "sent_timestamp": datetime.datetime.now().strftime("%m/%d/%Y %H:%M"),
                "needs_review": False,
            }
            data.setdefault("epo_log",[]).append(epo_entry)
            epo_log = data["epo_log"]
            persist()
            st.markdown(f'<div class="success-box">EPO for Lot {epo_lot} ({epo_comm}) added!</div>',unsafe_allow_html=True)

    # ââ EPO History ââ
    st.markdown("---")
    st.markdown("### EPO History")
    if not epo_log:
        st.info("No EPOs tracked yet. Send an EPO from Outlook with CC to the tracker email, then hit Sync.")
    else:
        filt = st.selectbox("Filter",["All","Pending","Confirmed","Denied","Follow-Up Needed","Discount Request","Needs Review"])
        for i, epo in enumerate(reversed(epo_log)):
            real_idx = len(epo_log)-1-i
            status = epo.get("status","Pending")
            days = _epo_days_open(epo)

            if filt == "Pending" and status != "Pending": continue
            if filt == "Confirmed" and status != "Confirmed": continue
            if filt == "Denied" and status != "Denied": continue
            if filt == "Discount Request" and status != "Discount Request": continue
            if filt == "Follow-Up Needed" and not (status == "Pending" and days >= 4): continue
            if filt == "Needs Review" and not epo.get("needs_review"): continue

            overdue_tag = " | FOLLOW UP!" if status == "Pending" and days >= 4 else ""
            review_tag = " | NEEDS REVIEW" if epo.get("needs_review") else ""
            lot_display = epo.get('lot','?')
            comm_display = epo.get('neighborhood', epo.get('community','?'))
            amt_display = epo.get('amount','?')

            with st.expander(f"Lot {lot_display} â {comm_display} | ${amt_display} | {status}{overdue_tag}{review_tag}"):
                # If needs review, show editable fields
                if epo.get("needs_review"):
                    st.markdown('<div style="background:#FFF2CC;padding:8px;border-radius:6px;margin-bottom:8px;font-size:13px;">'
                                'Could not fully parse this email. Please review and fix the fields below.</div>',unsafe_allow_html=True)
                    if epo.get("original_subject"):
                        st.caption(f"Original subject: {epo['original_subject']}")
                    with st.form(f"review_{real_idx}"):
                        r_comm = st.selectbox("Community", list(comms.keys()),
                            index=list(comms.keys()).index(epo.get("neighborhood","")) if epo.get("neighborhood","") in comms else 0,
                            key=f"rc_{real_idx}")
                        r_lot = st.text_input("Lot #", value=epo.get("lot",""), key=f"rl_{real_idx}")
                        r_amt = st.text_input("Amount ($)", value=epo.get("amount",""), key=f"ra_{real_idx}")
                        r_desc = st.text_input("Description", value=epo.get("description",""), key=f"rd_{real_idx}")
                        r_contact = st.text_input("Builder Contact", value=epo.get("builder_contact",""), key=f"rn_{real_idx}")
                        if st.form_submit_button("Save & Approve"):
                            epo_log[real_idx]["neighborhood"] = r_comm
                            epo_log[real_idx]["lot"] = r_lot
                            epo_log[real_idx]["amount"] = r_amt
                            epo_log[real_idx]["description"] = r_desc
                            epo_log[real_idx]["builder_contact"] = r_contact
                            epo_log[real_idx]["builder"] = comms.get(r_comm,{}).get("builder","")
                            epo_log[real_idx]["short_name"] = get_lot_code(r_comm, r_lot)
                            epo_log[real_idx]["needs_review"] = False
                            persist(); st.rerun()
                else:
                    st.write(f"**Builder Contact:** {epo.get('builder_contact','')}")
                    st.write(f"**Builder:** {epo.get('builder','')}")
                    st.write(f"**Date:** {epo.get('date','')}")
                    st.write(f"**Lot Code:** {epo.get('short_name','')}")
                    st.write(f"**Description:** {epo.get('description','')}")
                    st.write(f"**Amount:** ${epo.get('amount','')}")
                    st.write(f"**Days Open:** {days}")
                    if epo.get("confirmation_num"):
                        st.write(f"**Confirmation #:** {epo['confirmation_num']}")

                    # Status update buttons
                    if status == "Pending":
                        st.markdown("**Update Status:**")
                        col1,col2,col3,col4 = st.columns(4)
                        if col1.button("Confirmed",key=f"conf_{real_idx}"):
                            epo_log[real_idx]["status"] = "Confirmed"
                            persist(); st.rerun()
                        if col2.button("Denied",key=f"deny_{real_idx}"):
                            epo_log[real_idx]["status"] = "Denied"
                            persist(); st.rerun()
                        if col3.button("Discount Req",key=f"disc_{real_idx}"):
                            epo_log[real_idx]["status"] = "Discount Request"
                            persist(); st.rerun()
                        if col4.button("Follow Up",key=f"fu_{real_idx}"):
                            fu_subject = f"Follow-Up: EPO Request - Lot {epo.get('lot','')}, {epo.get('neighborhood','')}"
                            fu_body = (f"Hi {epo.get('builder_contact','')},\n\nFollowing up on the EPO request for "
                                       f"Lot {epo.get('lot','')} in {epo.get('neighborhood','')}.\n"
                                       f"Amount: ${epo.get('amount','')}\nOriginal request sent: {epo.get('date','')}\n\n"
                                       f"Please advise on the status and provide a confirmation number.\n\nThank you,\nGabriel Jordao")
                            email_to = epo.get("builder_contact_email","")
                            ok, fmsg = send_email(email_to, fu_subject, fu_body, cc=TRACKER_EMAIL)
                            epo_log[real_idx]["followup_count"] = epo.get("followup_count",0) + 1
                            persist()
                            if ok:
                                st.markdown(f'<div class="success-box">Follow-up email sent!</div>',unsafe_allow_html=True)
                            else:
                                st.warning(f"Could not send follow-up: {fmsg}")
                                st.code(f"To: {email_to}\nCC: {TRACKER_EMAIL}\nSubject: {fu_subject}\n\n{fu_body}", language=None)

                    # Add confirmation number
                    if status == "Confirmed" and not epo.get("confirmation_num"):
                        with st.form(f"confnum_{real_idx}"):
                            conf_num = st.text_input("Enter Confirmation/PO Number")
                            if st.form_submit_button("Save"):
                                epo_log[real_idx]["confirmation_num"] = conf_num
                                persist(); st.rerun()

    # ââ Export to Excel ââ
    st.markdown("---")
    if st.button("Export EPO Spreadsheet"):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            ewb = Workbook()
            ews = ewb.active
            ews.title = "Open EPOs"
            headers = ["FIELD MANAGER","BUILDER CONTACT","DATE","BUILDER","NEIGHBORHOOD",
                       "LOT","SHORT NAME","DESCRIPTION","AMOUNT","STATUS","CONFIRMATION #","DAYS OPEN"]
            hfont = Font(name="Aptos Narrow",bold=True,size=11)
            hfill = PatternFill(start_color="1F4E79",end_color="1F4E79",fill_type="solid")
            hfont_w = Font(name="Aptos Narrow",bold=True,size=11,color="FFFFFF")
            bfont = Font(name="Aptos Narrow",size=11)
            thin_border = Border(
                left=Side(style="thin"),right=Side(style="thin"),
                top=Side(style="thin"),bottom=Side(style="thin"))
            for c,h in enumerate(headers,1):
                cell = ews.cell(row=1,column=c,value=h)
                cell.font = hfont_w
                cell.fill = hfill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            widths = [18,20,12,12,18,8,14,45,12,16,16,12]
            for c,w in enumerate(widths,1):
                col_letter = chr(64+c) if c <= 26 else "A" + chr(64+c-26)
                ews.column_dimensions[col_letter].width = w
            # Status color fills
            status_fills = {
                "Confirmed": PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid"),
                "Pending": PatternFill(start_color="FFF2CC",end_color="FFF2CC",fill_type="solid"),
                "Denied": PatternFill(start_color="FFC7CE",end_color="FFC7CE",fill_type="solid"),
                "Discount Request": PatternFill(start_color="FCE4D6",end_color="FCE4D6",fill_type="solid"),
            }
            for r,e in enumerate(epo_log,2):
                ews.cell(row=r,column=1,value=e.get("field_manager",FIELD_MANAGER)).font=bfont
                ews.cell(row=r,column=2,value=e.get("builder_contact","")).font=bfont
                ews.cell(row=r,column=3,value=e.get("date","")).font=bfont
                ews.cell(row=r,column=4,value=e.get("builder","")).font=bfont
                ews.cell(row=r,column=5,value=e.get("neighborhood","")).font=bfont
                ews.cell(row=r,column=6,value=e.get("lot","")).font=bfont
                ews.cell(row=r,column=7,value=e.get("short_name","")).font=bfont
                ews.cell(row=r,column=8,value=e.get("description","")).font=bfont
                amt_str = e.get("amount","")
                try:
                    ews.cell(row=r,column=9,value=float(str(amt_str).replace(",","").replace("$",""))).font=bfont
                    ews.cell(row=r,column=9).number_format = "$#,##0.00"
                except:
                    ews.cell(row=r,column=9,value=amt_str).font=bfont
                status_val = e.get("status","Pending")
                status_cell = ews.cell(row=r,column=10,value=status_val)
                status_cell.font = bfont
                if status_val in status_fills:
                    status_cell.fill = status_fills[status_val]
                ews.cell(row=r,column=11,value=e.get("confirmation_num","")).font=bfont
                ews.cell(row=r,column=12,value=_epo_days_open(e)).font=bfont
                # Apply borders
                for col in range(1,13):
                    ews.cell(row=r,column=col).border = thin_border
            # Summary row
            summary_row = len(epo_log) + 3
            ews.cell(row=summary_row, column=1, value="SUMMARY").font = hfont
            ews.cell(row=summary_row+1, column=1, value="Total EPOs:").font = bfont
            ews.cell(row=summary_row+1, column=2, value=total_epos).font = bfont
            ews.cell(row=summary_row+2, column=1, value="Confirmed:").font = bfont
            ews.cell(row=summary_row+2, column=2, value=confirmed).font = bfont
            ews.cell(row=summary_row+3, column=1, value="Capture Rate:").font = bfont
            ews.cell(row=summary_row+3, column=2, value=capture_rate).font = bfont
            export_path = os.path.join(DATA_DIR, "EPO_Tracker.xlsx")
            ewb.save(export_path)
            with open(export_path, "rb") as f:
                st.download_button("Download EPO Spreadsheet", f.read(), "EPO_Tracker.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as ex:
            st.error(f"Export failed: {ex}")

# âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# FIELD NOTES
# âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
elif page == "Field Notes":
    st.markdown('<div class="page-title">Field Notes</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-subtitle">Quick notes from the field with smart actions</div>', unsafe_allow_html=True)
    data = get_data()
    comms = data["communities"]
    st.caption("Quick notes from the field. Type lot # + note, one per line.")
    comm_name = st.selectbox("Community", list(comms.keys()), key="notes_comm")

    with st.form("notes_form",clear_on_submit=True):
        raw = st.text_area("Enter notes (LOT# your note)",height=150,
            placeholder="101 needs touch up on master bedroom\n102 ready for final paint\n103 schedule clean-out")
        parse_btn = st.form_submit_button("Parse Notes")

    if parse_btn and raw:
        parsed = []
        for line in raw.strip().splitlines():
            line = line.strip()
            if not line: continue
            parts = line.split(" ",1)
            lot_n, note_text = parts[0], parts[1] if len(parts)>1 else ""
            lower = note_text.lower()
            action = "Note Logged"
            if any(k in lower for k in ("clean-out","clean out","schedule clean")): action="Schedule Clean-Out"
            elif any(k in lower for k in ("ready for final","final paint","final point")): action="Notify for Final"
            elif "epo" in lower: action="Request EPO"
            elif any(k in lower for k in ("touch up","touchup","repair")): action="Schedule Repair"
            elif any(k in lower for k in ("ready for qc","qc ready")): action="Schedule QC"
            elif any(k in lower for k in ("homeowner","ho walk","ho ready")): action="Schedule Homeowner Walk"
            elif any(k in lower for k in ("hang ready","ready for hang","framing done")): action="Schedule Hang"
            parsed.append({"lot":lot_n,"community":comm_name,"note":note_text,"action":action,
                "timestamp":datetime.datetime.now().strftime("%m/%d/%Y %H:%M")})
            if lot_n in comms.get(comm_name,{}).get("lots",{}):
                existing = comms[comm_name]["lots"][lot_n].get("notes","")
                comms[comm_name]["lots"][lot_n]["notes"] = (existing+" | "+note_text).strip(" | ")
                # Also store structured note history per lot
                lot_notes = comms[comm_name]["lots"][lot_n].setdefault("note_history", [])
                lot_notes.append({"text": note_text, "action": action, "timestamp": datetime.datetime.now().strftime("%m/%d/%Y %H:%M")})
        data.setdefault("notes",[]).extend(parsed); persist()

        st.markdown("### Parsed Notes")
        colors = {"Schedule Clean-Out":"#FCE4D6","Notify for Final":"#BDD7EE","Request EPO":"#FFF2CC",
            "Schedule Repair":"#FCE4D6","Schedule QC":"#E2EFDA","Schedule Homeowner Walk":"#E2EFDA","Schedule Hang":"#C6EFCE","Note Logged":"#f0f0f0"}
        for p in parsed:
            c = colors.get(p["action"],"#f0f0f0")
            st.markdown(f'<div style="padding:10px;margin:6px 0;border-radius:8px;background:{c};border-left:4px solid #1F4E79;"><strong>Lot {p["lot"]}</strong> â {p["note"]}<br><small>Action: <strong>{p["action"]}</strong></small></div>',unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### Recent Notes")
    all_notes = data.get("notes",[])
    if all_notes:
        for note in reversed(all_notes[-20:]):
            st.text(f"[{note.get('timestamp','')}] Lot {note['lot']} ({note['community']}): {note['note']} â {note['action']}")
    else:
        st.info("No notes yet.")

# âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# SETTINGS
# âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
elif page == "Settings":
    st.markdown('<div class="page-title">Settings</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-subtitle">Email config, communities, and subs</div>', unsafe_allow_html=True)
    data = get_data()
    comms = data["communities"]
    cfg = st.session_state.config

    # Email setup
    st.markdown("### Email Configuration")
    st.markdown(f"**Tracker Email:** {TRACKER_EMAIL}")
    st.markdown(f"**Drywall Orders Email:** {DW_ORDERS_EMAIL}")
    with st.form("email_config"):
        app_pw = st.text_input("Gmail App Password", value=cfg.get("gmail_app_password",""), type="password",
            help="Generate this at myaccount.google.com â Security â App Passwords")
        if st.form_submit_button("Save Email Config"):
            cfg["gmail_app_password"] = app_pw
            save_config(cfg)
            st.session_state.config = cfg
            st.success("Email config saved!")
            # Test connection
            try:
                with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
                    server.login(TRACKER_EMAIL, app_pw)
                st.markdown('<div class="success-box">Connection test passed! Emails will work.</div>',unsafe_allow_html=True)
            except Exception as e:
                st.error(f"Connection test failed: {e}")

    st.markdown("---")
    st.markdown("### Community Settings")
    comm_name = st.selectbox("Select Community", list(comms.keys()), key="settings_comm")
    comm = comms[comm_name]

    # Lot code preview
    code_cfg = LOT_CODES.get(comm_name)
    if code_cfg:
        sample_lot = list(comm["lots"].keys())[0] if comm["lots"] else "1"
        st.markdown(f"**Lot Code Format:** {get_lot_code(comm_name, sample_lot)}")

    # Durations
    st.markdown("#### Task Durations (days)")
    durations = comm.get("durations",{"Hang":1,"Scrap":1,"Tape":2,"Bed":2,"Skim":2,"Sand":1})
    with st.form(f"dur_{comm_name}"):
        new_dur = {}
        cols = st.columns(3)
        for i, task in enumerate(DRYWALL_STAGES):
            new_dur[task] = cols[i%3].number_input(task,min_value=1,max_value=10,value=durations.get(task,1),key=f"d_{comm_name}_{task}")
        if st.form_submit_button("Save Durations"):
            comm["durations"]=new_dur; persist(); st.success("Saved!")

    # Subs
    st.markdown("#### Drywall Subs")
    subs = comm.get("subs",{})
    with st.form(f"sub_{comm_name}"):
        new_subs = {}
        for task in DRYWALL_STAGES:
            new_subs[task] = st.text_input(task,value=subs.get(task,"TBD"),key=f"sb_{comm_name}_{task}")
        if st.form_submit_button("Save Subs"):
            comm["subs"]=new_subs; persist(); st.success("Saved!")

    st.markdown("---")
    if st.button("Export All Data as JSON"):
        st.download_button("Download",json.dumps(data,indent=2,default=str),"stancil_field_data.json","application/json")

# Footer
st.markdown('<div class="sf-footer">Stancil Field Manager v2.0</div>', unsafe_allow_html=True)
